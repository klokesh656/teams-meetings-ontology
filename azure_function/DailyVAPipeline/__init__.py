"""
Azure Function: Daily VA Check-in Pipeline
==========================================
Timer-triggered function that runs daily at 6:00 AM UTC.

This function:
1. Downloads new meeting transcripts from Microsoft Graph API
2. Analyzes check-in transcripts with Churn Risk Checklist
3. Generates CSV files for Power BI/Power Automate
4. Uploads results to Azure Blob Storage
"""

import azure.functions as func
import logging
import os
import json
import re
import csv
import io
from datetime import datetime, timedelta
from urllib.parse import quote
import requests

from azure.identity import ClientSecretCredential
from azure.storage.blob import BlobServiceClient

# ============================================================================
# CONFIGURATION (from Azure Function App Settings)
# ============================================================================

TENANT_ID = os.environ.get('AZURE_TENANT_ID')
CLIENT_ID = os.environ.get('AZURE_CLIENT_ID')
CLIENT_SECRET = os.environ.get('AZURE_CLIENT_SECRET')
HR_USER_ID = os.environ.get('HR_USER_ID', '81835016-79d5-4a15-91b1-c104e2cd9adb')

STORAGE_CONNECTION_STRING = os.environ.get('AZURE_STORAGE_CONNECTION_STRING')
STORAGE_ACCOUNT = os.environ.get('STORAGE_ACCOUNT', 'aidevelopement')
TRANSCRIPT_CONTAINER = os.environ.get('TRANSCRIPT_CONTAINER', 'transcripts')
OUTPUT_CONTAINER = os.environ.get('OUTPUT_CONTAINER', 'pipeline-outputs')

AZURE_OPENAI_ENDPOINT = os.environ.get('AZURE_OPENAI_ENDPOINT')
AZURE_OPENAI_KEY = os.environ.get('AZURE_OPENAI_KEY')
AZURE_OPENAI_DEPLOYMENT = os.environ.get('AZURE_OPENAI_DEPLOYMENT', 'gpt-4.1')

DAYS_TO_LOOK_BACK = int(os.environ.get('DAYS_TO_LOOK_BACK', '7'))

# SharePoint Configuration - BT Group plc site
SHAREPOINT_SITE_ID = os.environ.get('SHAREPOINT_SITE_ID', 'bt2685218p1.sharepoint.com,8470aef2-afc9-4665-8730-e63c87b0ebac,66af7ccb-d9d6-41da-b647-1c62bc4fd46e')
SHAREPOINT_DRIVE_ID = os.environ.get('SHAREPOINT_DRIVE_ID', 'b!8q5whMmvZUaHMOY8h7DrrMt8r2bW2dpBtkccYrxP1G6UufIavqAUT4u5gAnnEB1q')
SHAREPOINT_FOLDER = os.environ.get('SHAREPOINT_FOLDER', 'Client Feedback')

# Azure AI Search Configuration - for Copilot Agent indexing
AZURE_SEARCH_ENDPOINT = os.environ.get('AZURE_SEARCH_ENDPOINT', 'https://aisearch-lokesg-2.search.windows.net')
AZURE_SEARCH_KEY = os.environ.get('AZURE_SEARCH_KEY', '')
AZURE_SEARCH_INDEX = os.environ.get('AZURE_SEARCH_INDEX', 'copilot-agent-comprehensive')
AZURE_SEARCH_COACH_INDEX = os.environ.get('AZURE_SEARCH_COACH_INDEX', 'coach-performance-index')

# ============================================================================
# CHURN RISK CHECKLIST
# ============================================================================

CHURN_RISK_CHECKLIST = """
## VA-Side Signals (VA001-VA012)
- VA001: Resignation hints or job search mentions
- VA002: Attendance issues
- VA003: Health concerns
- VA004: Overwhelming workload/burnout
- VA005: Skill gaps
- VA006: Communication breakdowns
- VA007: Disengagement
- VA008: Personal issues
- VA009: Compensation concerns
- VA010: Career growth frustration
- VA011: Work-life balance issues
- VA012: Tool/system problems

## Client-Side Signals (CL001-CL010)
- CL001: Negative feedback
- CL002: Scope creep
- CL003: Non-responsive client
- CL004: Payment issues
- CL005: Micromanagement
- CL006: Budget cuts
- CL007: Organizational changes
- CL008: Multiple VA requests
- CL009: Reduced hours
- CL010: Communication mismatch

## Relationship/HR Signals (RH001-RH005)
- RH001: VA-Client conflict
- RH002: Integration escalations
- RH003: Contract at risk
- RH004: Missing check-ins
- RH005: No visibility
"""

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_graph_headers():
    """Get Microsoft Graph API headers."""
    credential = ClientSecretCredential(TENANT_ID, CLIENT_ID, CLIENT_SECRET)
    token = credential.get_token('https://graph.microsoft.com/.default')
    return {
        'Authorization': f'Bearer {token.token}',
        'Content-Type': 'application/json'
    }


def upload_to_sharepoint(file_content, filename, folder_path=None, logging=None):
    """Upload file to SharePoint using app-only authentication."""
    try:
        # Get access token
        token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
        token_data = {
            'grant_type': 'client_credentials',
            'client_id': CLIENT_ID,
            'client_secret': CLIENT_SECRET,
            'scope': 'https://graph.microsoft.com/.default'
        }
        token_response = requests.post(token_url, data=token_data)
        access_token = token_response.json().get('access_token')
        
        if not access_token:
            if logging:
                logging.warning("   ⚠️ Failed to get SharePoint access token")
            return None
        
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/octet-stream'
        }
        
        folder = folder_path or SHAREPOINT_FOLDER
        
        # Use site-specific endpoint for app-only auth
        upload_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}/root:/{folder}/{filename}:/content"
        
        response = requests.put(upload_url, headers=headers, data=file_content)
        
        if response.status_code in [200, 201]:
            web_url = response.json().get('webUrl', '')
            if logging:
                logging.info(f"   ✅ Uploaded to SharePoint: {folder}/{filename}")
            return web_url
        else:
            if logging:
                logging.warning(f"   ⚠️ SharePoint upload returned {response.status_code}")
            return None
            
    except Exception as e:
        if logging:
            logging.warning(f"   ⚠️ SharePoint upload failed: {e}")
        return None


def get_blob_service():
    """Get Azure Blob Storage client."""
    return BlobServiceClient.from_connection_string(STORAGE_CONNECTION_STRING)


def generate_blob_url(filename):
    """Generate blob URL."""
    encoded = quote(filename, safe='')
    return f"https://{STORAGE_ACCOUNT}.blob.core.windows.net/{TRANSCRIPT_CONTAINER}/{encoded}"


def generate_meeting_id(source_file, va_name='', meeting_date=''):
    """Generate a consistent Meeting ID based on source file.
    
    Creates a unique, stable ID that remains the same across pipeline runs,
    enabling Power BI relationships and Power Automate status tracking.
    """
    import hashlib
    if source_file:
        file_hash = hashlib.md5(source_file.encode()).hexdigest()[:6].upper()
        return f"MTG-{file_hash}"
    elif va_name and meeting_date:
        combined = f"{va_name}_{meeting_date}"
        file_hash = hashlib.md5(combined.encode()).hexdigest()[:6].upper()
        return f"MTG-{file_hash}"
    return "MTG-UNKNOWN"


def load_review_status(blob_service):
    """Load review status from blob storage."""
    try:
        container = blob_service.get_container_client(OUTPUT_CONTAINER)
        blob = container.get_blob_client("data/meeting_review_status.json")
        data = blob.download_blob().readall()
        return json.loads(data)
    except:
        return {}


def load_state(blob_service):
    """Load pipeline state from blob storage."""
    try:
        container = blob_service.get_container_client(OUTPUT_CONTAINER)
        blob = container.get_blob_client("state/pipeline_state.json")
        data = blob.download_blob().readall()
        return json.loads(data)
    except:
        return {'processed_files': [], 'last_run': None}


def save_state(blob_service, state):
    """Save pipeline state to blob storage."""
    container = blob_service.get_container_client(OUTPUT_CONTAINER)
    blob = container.get_blob_client("state/pipeline_state.json")
    blob.upload_blob(json.dumps(state, indent=2, default=str), overwrite=True)


def load_history(blob_service):
    """Load analysis history from blob storage."""
    try:
        container = blob_service.get_container_client(OUTPUT_CONTAINER)
        blob = container.get_blob_client("data/checkin_analysis_history.json")
        data = blob.download_blob().readall()
        return json.loads(data)
    except:
        return {'analyses': [], 'last_updated': None}


def save_history(blob_service, history):
    """Save analysis history to blob storage."""
    container = blob_service.get_container_client(OUTPUT_CONTAINER)
    blob = container.get_blob_client("data/checkin_analysis_history.json")
    blob.upload_blob(json.dumps(history, indent=2, default=str), overwrite=True)


# ============================================================================
# PIPELINE STEPS
# ============================================================================

def download_new_transcripts(logging, days_back):
    """Download new transcripts from Graph API."""
    logging.info(f"📥 Checking for transcripts (last {days_back} days)")
    
    headers = get_graph_headers()
    blob_service = get_blob_service()
    
    # Get existing transcripts from blob
    container = blob_service.get_container_client(TRANSCRIPT_CONTAINER)
    try:
        container.create_container()
    except:
        pass
    
    existing = set()
    for blob in container.list_blobs():
        match = re.search(r'(\d{8}_\d{4})', blob.name)
        if match:
            existing.add(match.group(1))
    
    # Fetch from Graph API
    url = f"https://graph.microsoft.com/beta/users/{HR_USER_ID}/onlineMeetings/getAllTranscripts(meetingOrganizerUserId='{HR_USER_ID}')"
    
    all_transcripts = []
    while url:
        resp = requests.get(url, headers=headers, timeout=60)
        if resp.status_code != 200:
            break
        data = resp.json()
        all_transcripts.extend(data.get('value', []))
        url = data.get('@odata.nextLink')
    
    # Filter recent
    cutoff = (datetime.utcnow() - timedelta(days=days_back)).isoformat() + 'Z'
    recent = [t for t in all_transcripts if t.get('createdDateTime', '') >= cutoff]
    
    logging.info(f"   Found {len(recent)} recent transcripts")
    
    # Download new ones
    downloaded = []
    for t in recent:
        created = t.get('createdDateTime', '')[:19].replace('T', ' ')
        date_str = created[:10].replace('-', '')
        time_str = created[11:16].replace(':', '')
        key = f"{date_str}_{time_str}"
        
        if key in existing:
            continue
        
        meeting_id = t.get('meetingId', '')
        transcript_id = t.get('id', '')
        
        # Get subject
        subject_url = f"https://graph.microsoft.com/beta/users/{HR_USER_ID}/onlineMeetings/{meeting_id}"
        try:
            resp = requests.get(subject_url, headers=headers, timeout=30)
            subject = resp.json().get('subject', 'Unknown') if resp.status_code == 200 else 'Unknown'
        except:
            subject = 'Unknown'
        
        # Download content
        content_url = f"https://graph.microsoft.com/beta/users/{HR_USER_ID}/onlineMeetings/{meeting_id}/transcripts/{transcript_id}/content"
        try:
            resp = requests.get(content_url, headers=headers, params={'$format': 'text/vtt'}, timeout=60)
            if resp.status_code == 200:
                content = resp.text
                
                safe_subject = re.sub(r'[<>:"/\\|?*]', '', subject)[:60]
                filename = f"{date_str}_{time_str}_{safe_subject}.vtt"
                
                # Upload to blob
                blob_client = container.get_blob_client(filename)
                blob_client.upload_blob(content, overwrite=True)
                
                downloaded.append({
                    'filename': filename,
                    'date': f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}",
                    'subject': subject,
                    'content': content
                })
                
                logging.info(f"   ✅ {filename}")
        except Exception as e:
            logging.warning(f"   Failed: {e}")
    
    logging.info(f"   Downloaded {len(downloaded)} new transcripts")
    return downloaded


def analyze_transcript(content, va_name, date_str, logging):
    """Analyze a single transcript with Azure OpenAI."""
    prompt = f"""Analyze this VA check-in meeting transcript for churn risk.

CHECKLIST:
{CHURN_RISK_CHECKLIST}

TRANSCRIPT:
{content[:12000]}

JSON response:
{{
    "va_name": "{va_name}",
    "client_name": "extracted or Unknown",
    "meeting_date": "{date_str}",
    "overall_risk_level": "low|medium|high|critical",
    "va_status": "green|yellow|red",
    "client_health": "healthy|at_risk|critical",
    "detected_signals": [{{"signal_id": "...", "evidence": "...", "confidence": "..."}}],
    "executive_summary": "2-3 sentences",
    "key_findings": ["..."],
    "ai_suggestions": [{{"issue": "...", "suggestion": "...", "urgency": "...", "category": "..."}}]
}}"""

    try:
        headers = {'Content-Type': 'application/json', 'api-key': AZURE_OPENAI_KEY}
        payload = {
            'messages': [
                {'role': 'system', 'content': 'Expert HR analyst for VA retention.'},
                {'role': 'user', 'content': prompt}
            ],
            'temperature': 0.3,
            'max_tokens': 2000
        }
        
        url = f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/{AZURE_OPENAI_DEPLOYMENT}/chat/completions?api-version=2024-02-15-preview"
        resp = requests.post(url, headers=headers, json=payload, timeout=120)
        
        if resp.status_code == 200:
            result = resp.json()['choices'][0]['message']['content']
            match = re.search(r'\{[\s\S]*\}', result)
            if match:
                return json.loads(match.group())
    except Exception as e:
        logging.error(f"Analysis error: {e}")
    
    return None


def analyze_new_transcripts(downloaded, logging):
    """Analyze downloaded check-in transcripts."""
    logging.info("🔍 Analyzing new check-ins...")
    
    blob_service = get_blob_service()
    history = load_history(blob_service)
    
    analyzed = []
    for item in downloaded:
        if 'check-in' not in item['subject'].lower() and 'check in' not in item['subject'].lower():
            continue
        
        # Extract VA name
        va_name = None
        if ' x ' in item['filename'].lower():
            parts = item['filename'].split(' x ')
            if len(parts) >= 2:
                va_name = parts[-1].replace('.vtt', '').split('-')[0].strip()
        
        if not va_name:
            continue
        
        analysis = analyze_transcript(item['content'], va_name, item['date'], logging)
        if analysis:
            analysis['source_file'] = item['filename']
            analysis['analyzed_at'] = datetime.utcnow().isoformat()
            history['analyses'].append(analysis)
            analyzed.append(analysis)
            logging.info(f"   ✅ {va_name}: {analysis['overall_risk_level'].upper()}")
    
    history['last_updated'] = datetime.utcnow().isoformat()
    save_history(blob_service, history)
    
    logging.info(f"   Analyzed {len(analyzed)} check-ins")
    return analyzed


def generate_and_upload_csvs(logging):
    """Generate CSV files and upload to blob storage."""
    logging.info("📊 Generating CSV files...")
    
    blob_service = get_blob_service()
    history = load_history(blob_service)
    analyses = history.get('analyses', [])
    review_status = load_review_status(blob_service)
    
    container = blob_service.get_container_client(OUTPUT_CONTAINER)
    try:
        container.create_container()
    except:
        pass
    
    date_prefix = datetime.utcnow().strftime('%Y%m%d')
    
    # Generate VA Risk Summary
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Latest Meeting ID', 'VA Name', 'Total Check-ins', 'Reviewed', 'Pending',
                     'Critical', 'High', 'Medium', 'Low', 'Risk Score', 'Current Risk', 
                     'Latest Date', 'Blob Link'])
    
    va_stats = {}
    for a in analyses:
        va = a.get('va_name', 'Unknown')
        meeting_id = generate_meeting_id(a.get('source_file', ''), va, a.get('meeting_date', ''))
        
        if va not in va_stats:
            va_stats[va] = {'name': va, 'total': 0, 'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 
                          'reviewed': 0, 'pending': 0, 'latest_date': '', 'latest_risk': '', 
                          'latest_source': '', 'latest_meeting_id': ''}
        
        va_stats[va]['total'] += 1
        risk = a.get('overall_risk_level', 'medium').lower()
        va_stats[va][risk] = va_stats[va].get(risk, 0) + 1
        
        # Track review status
        status = review_status.get(meeting_id, {}).get('status', 'Pending')
        if status == 'Reviewed':
            va_stats[va]['reviewed'] += 1
        else:
            va_stats[va]['pending'] += 1
        
        date = a.get('meeting_date', '')
        if date >= va_stats[va]['latest_date']:
            va_stats[va]['latest_date'] = date
            va_stats[va]['latest_risk'] = risk
            va_stats[va]['latest_source'] = a.get('source_file', '')
            va_stats[va]['latest_meeting_id'] = meeting_id
    
    for va, s in sorted(va_stats.items(), key=lambda x: x[1]['critical']*100 + x[1]['high']*10, reverse=True):
        score = s['critical']*100 + s['high']*25 + s['medium']*5 + s['low']
        writer.writerow([s['latest_meeting_id'], va, s['total'], s['reviewed'], s['pending'],
                        s['critical'], s['high'], s['medium'], s['low'],
                        score, s['latest_risk'].upper(), s['latest_date'], 
                        generate_blob_url(s['latest_source'])])
    
    blob = container.get_blob_client(f"csv/{date_prefix}/va_risk_summary.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    blob = container.get_blob_client("latest/va_risk_summary.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    
    # Generate Critical Alerts
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Meeting ID', 'Priority', 'VA Name', 'Client', 'Date', 'Risk', 
                     'Review Status', 'Summary', 'Top Suggestion', 'Client Input', 'Blob Link'])
    
    for a in sorted(analyses, key=lambda x: x.get('meeting_date',''), reverse=True):
        risk = a.get('overall_risk_level', '').lower()
        if risk in ['critical', 'high']:
            meeting_id = generate_meeting_id(a.get('source_file', ''), a.get('va_name', ''), a.get('meeting_date', ''))
            priority = 'P1' if risk == 'critical' else 'P2'
            status_info = review_status.get(meeting_id, {})
            suggestions = a.get('ai_suggestions', [])
            top_suggestion = suggestions[0].get('suggestion', '')[:150] if suggestions else ''
            
            writer.writerow([
                meeting_id, priority, a.get('va_name', ''), a.get('client_name', 'Unknown'),
                a.get('meeting_date', ''), risk.upper(),
                status_info.get('status', 'Pending'),
                a.get('executive_summary', '')[:200],
                top_suggestion,
                status_info.get('client_input', ''),
                generate_blob_url(a.get('source_file', ''))
            ])
    
    blob = container.get_blob_client(f"csv/{date_prefix}/critical_alerts.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    blob = container.get_blob_client("latest/critical_alerts.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    
    # Generate All Meetings
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Meeting ID', 'VA Name', 'Client', 'Date', 'Risk', 'Review Status',
                     'Signals', 'Summary', 'Top Suggestions', 'Client Input', 'Blob Link'])
    
    for a in sorted(analyses, key=lambda x: x.get('meeting_date',''), reverse=True):
        meeting_id = generate_meeting_id(a.get('source_file', ''), a.get('va_name', ''), a.get('meeting_date', ''))
        status_info = review_status.get(meeting_id, {})
        suggestions = a.get('ai_suggestions', [])
        top_suggestions = '; '.join([s.get('suggestion', '')[:100] for s in suggestions[:2]])
        
        writer.writerow([
            meeting_id, a.get('va_name', ''), a.get('client_name', 'Unknown'),
            a.get('meeting_date', ''), a.get('overall_risk_level', '').upper(),
            status_info.get('status', 'Pending'),
            len(a.get('detected_signals', [])), a.get('executive_summary', '')[:300],
            top_suggestions[:200],
            status_info.get('client_input', ''),
            generate_blob_url(a.get('source_file', ''))
        ])
    
    blob = container.get_blob_client(f"csv/{date_prefix}/all_meetings_detail.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    blob = container.get_blob_client("latest/all_meetings_detail.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    
    # Generate KPI Summary
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['KPI', 'Value', 'Updated'])
    
    total = len(analyses)
    critical = sum(1 for a in analyses if a.get('overall_risk_level','').lower() == 'critical')
    high = sum(1 for a in analyses if a.get('overall_risk_level','').lower() == 'high')
    reviewed = sum(1 for a in analyses if review_status.get(generate_meeting_id(a.get('source_file','')), {}).get('status') == 'Reviewed')
    
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M')
    writer.writerow(['Total VAs', len(va_stats), now])
    writer.writerow(['Total Check-ins', total, now])
    writer.writerow(['Critical Risk', critical, now])
    writer.writerow(['High Risk', high, now])
    writer.writerow(['Reviewed', reviewed, now])
    writer.writerow(['Pending Review', total - reviewed, now])
    writer.writerow(['At-Risk Rate %', f'{(critical+high)/total*100:.1f}' if total else '0', now])
    
    blob = container.get_blob_client(f"csv/{date_prefix}/kpi_dashboard_summary.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    blob = container.get_blob_client("latest/kpi_dashboard_summary.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    
    logging.info("   ✅ CSV files uploaded")
    
    # Generate Client Feedback Form (XLSX with Excel Table)
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        
        headers = ['Meeting_ID', 'VA_Name', 'Client', 'Meeting_Date', 'Risk_Level',
                   'Issue_Summary', 'AI_Recommendation', 'Client_Agrees', 'Client_Priority',
                   'Client_Suggestion', 'Client_Action_Taken', 'Client_Notes', 
                   'Reviewer_Name', 'Review_Date', 'Feedback_Status']
        
        rows = []
        for a in analyses:
            risk = a.get('overall_risk_level', '').lower()
            if risk in ['critical', 'high', 'medium']:
                meeting_id = generate_meeting_id(a.get('source_file', ''), a.get('va_name', ''), a.get('meeting_date', ''))
                suggestions = a.get('ai_suggestions', [])
                top_suggestion = suggestions[0].get('suggestion', '')[:300] if suggestions else ''
                
                rows.append({
                    'Meeting_ID': meeting_id,
                    'VA_Name': a.get('va_name', ''),
                    'Client': (a.get('client_name', '') or 'Unknown')[:50],
                    'Meeting_Date': a.get('meeting_date', ''),
                    'Risk_Level': risk.upper(),
                    'Issue_Summary': a.get('executive_summary', '')[:300],
                    'AI_Recommendation': top_suggestion,
                    'Client_Agrees': '',
                    'Client_Priority': '',
                    'Client_Suggestion': '',
                    'Client_Action_Taken': '',
                    'Client_Notes': '',
                    'Reviewer_Name': '',
                    'Review_Date': '',
                    'Feedback_Status': 'PENDING'
                })
        
        if rows:
            wb = Workbook()
            ws = wb.active
            ws.title = "ClientFeedback"
            
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            for row_num, row_data in enumerate(rows, 2):
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
            
            table_ref = f"A1:O{len(rows) + 1}"
            table = Table(displayName="ClientFeedbackTable", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
            
            xlsx_output = io.BytesIO()
            wb.save(xlsx_output)
            xlsx_output.seek(0)
            
            blob = container.get_blob_client(f"csv/{date_prefix}/client_feedback_form.xlsx")
            blob.upload_blob(xlsx_output.getvalue(), overwrite=True)
            blob = container.get_blob_client("latest/client_feedback_form.xlsx")
            xlsx_output.seek(0)
            blob.upload_blob(xlsx_output.getvalue(), overwrite=True)
            logging.info(f"   ✅ client_feedback_form.xlsx ({len(rows)} items)")
            
            # Upload to SharePoint
            xlsx_output.seek(0)
            sp_url = upload_to_sharepoint(xlsx_output.getvalue(), 'client_feedback_form.xlsx', logging=logging)
            if sp_url:
                logging.info(f"   📤 SharePoint: {sp_url}")
            
    except ImportError:
        logging.warning("   ⚠️ openpyxl not available, skipping XLSX generation")
    except Exception as e:
        logging.warning(f"   ⚠️ XLSX generation failed: {e}")
    
    # Generate Coach Performance CSVs
    known_coaches = ['Louise', 'Shey', 'Shey Geraldes']
    coach_stats = {}
    coach_meetings = {}
    
    for analysis in analyses:
        coach_name = None
        coach_evidence = []
        
        for signal in analysis.get('detected_signals', []):
            evidence = signal.get('evidence', '')
            for coach in known_coaches:
                if coach.lower() in evidence.lower():
                    coach_name = 'Shey Geraldes' if 'shey' in coach.lower() else coach
                    coach_evidence.append(evidence)
                    break
        
        for finding in analysis.get('key_findings', []):
            for coach in known_coaches:
                if coach.lower() in finding.lower():
                    coach_name = 'Shey Geraldes' if 'shey' in coach.lower() else coach
                    coach_evidence.append(finding)
        
        for positive in analysis.get('positive_indicators', []):
            for coach in known_coaches:
                if coach.lower() in positive.lower():
                    coach_name = 'Shey Geraldes' if 'shey' in coach.lower() else coach
                    coach_evidence.append(f"[POSITIVE] {positive}")
        
        if coach_name:
            if coach_name not in coach_stats:
                coach_stats[coach_name] = {'total': 0, 'critical': 0, 'high': 0, 'medium': 0, 'low': 0,
                                           'escalations': 0, 'positive': 0, 'negative': 0, 'vas': set()}
                coach_meetings[coach_name] = []
            
            s = coach_stats[coach_name]
            s['total'] += 1
            s['vas'].add(analysis.get('va_name', 'Unknown'))
            
            risk = analysis.get('overall_risk_level', 'low').lower()
            s[risk] = s.get(risk, 0) + 1
            
            if analysis.get('escalation_needed'):
                s['escalations'] += 1
            
            for ev in coach_evidence:
                if '[POSITIVE]' in ev or 'support' in ev.lower() or 'help' in ev.lower():
                    s['positive'] += 1
                elif 'issue' in ev.lower() or 'problem' in ev.lower():
                    s['negative'] += 1
            
            meeting_id = generate_meeting_id(analysis.get('source_file', ''), analysis.get('va_name', ''), analysis.get('meeting_date', ''))
            coach_meetings[coach_name].append({
                'meeting_id': meeting_id, 'va': analysis.get('va_name', ''),
                'client': (analysis.get('client_name', '') or 'Unknown')[:50],
                'date': analysis.get('meeting_date', ''), 'risk': risk.upper(),
                'evidence': ' | '.join(coach_evidence[:2])[:400]
            })
    
    # Upload coach summary
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Coach Name', 'Total Meetings', 'VAs Coached', 'Critical', 'High', 'Medium', 'Low',
                     'Escalations', 'Positive Mentions', 'Negative Mentions', 'Prof Score', 'Rating'])
    
    for coach, s in coach_stats.items():
        total_mentions = s['positive'] + s['negative']
        prof_score = int((s['positive'] / total_mentions) * 100) if total_mentions > 0 else 50
        rating = 'EXCELLENT' if prof_score >= 80 and s['escalations'] > 0 else ('GOOD' if prof_score >= 60 else 'NEEDS REVIEW')
        
        writer.writerow([coach, s['total'], len(s['vas']), s['critical'], s['high'], s['medium'], s['low'],
                        s['escalations'], s['positive'], s['negative'], prof_score, rating])
    
    blob = container.get_blob_client(f"csv/{date_prefix}/coach_performance_summary.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    blob = container.get_blob_client("latest/coach_performance_summary.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    
    # Upload coach detail
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Meeting ID', 'Coach Name', 'VA Name', 'Client', 'Date', 'Risk', 'Evidence', 'Notes'])
    
    for coach, meetings in coach_meetings.items():
        for m in meetings:
            writer.writerow([m['meeting_id'], coach, m['va'], m['client'], m['date'], m['risk'], m['evidence'], ''])
    
    blob = container.get_blob_client(f"csv/{date_prefix}/coach_performance_detail.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    blob = container.get_blob_client("latest/coach_performance_detail.csv")
    blob.upload_blob(output.getvalue(), overwrite=True)
    
    logging.info(f"   ✅ Coach performance CSVs ({len(coach_stats)} coaches)")
    
    # Return data for indexing
    return {
        'analyses': analyses,
        'coach_stats': coach_stats,
        'coach_meetings': coach_meetings
    }


def update_copilot_search_index(data, logging):
    """Update Azure AI Search index with latest data for Copilot Agent."""
    if not AZURE_SEARCH_KEY:
        logging.warning("   ⚠️ Azure Search not configured, skipping index update")
        return
    
    logging.info("🔍 Updating Azure AI Search indexes for Copilot...")
    
    try:
        from azure.core.credentials import AzureKeyCredential
        from azure.search.documents import SearchClient
        from azure.search.documents.indexes import SearchIndexClient
        from azure.search.documents.indexes.models import (
            SearchIndex, SearchField, SearchFieldDataType,
            SimpleField, SearchableField,
            SemanticConfiguration, SemanticField, SemanticPrioritizedFields,
            SemanticSearch
        )
        
        credential = AzureKeyCredential(AZURE_SEARCH_KEY)
        index_client = SearchIndexClient(endpoint=AZURE_SEARCH_ENDPOINT, credential=credential)
        
        # =====================================================================
        # Update main Copilot index with meeting analyses
        # =====================================================================
        analyses = data.get('analyses', [])
        if analyses:
            documents = []
            now_iso = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
            
            for a in analyses:
                meeting_id = generate_meeting_id(a.get('source_file', ''), a.get('va_name', ''), a.get('meeting_date', ''))
                
                # Extract signals as list
                signals = [s.get('signal_id', '') for s in a.get('detected_signals', [])]
                suggestions = [s.get('suggestion', '')[:200] for s in a.get('ai_suggestions', [])[:3]]
                findings = [f[:200] for f in a.get('key_findings', [])[:3]]
                positives = [p[:200] for p in a.get('positive_indicators', [])[:3]]
                
                doc = {
                    'id': meeting_id,
                    'document_type': 'meeting_analysis',
                    'entity_type': 'va_checkin',
                    'title': f"Check-in: {a.get('va_name', 'Unknown')} - {a.get('meeting_date', '')}",
                    'content': a.get('executive_summary', '')[:2000],
                    'summary': a.get('executive_summary', '')[:500],
                    'va_name': a.get('va_name', 'Unknown'),
                    'client_name': (a.get('client_name', '') or 'Unknown')[:100],
                    'date_string': a.get('meeting_date', ''),
                    'risk_level': a.get('overall_risk_level', 'medium').upper(),
                    'risk_score': int(a.get('risk_score', 0)) if a.get('risk_score') else 0,
                    'health_status': a.get('client_health', 'unknown'),
                    'status': 'analyzed',
                    'risk_signals': signals,
                    'suggestions': suggestions,
                    'key_insights': findings,
                    'key_positives': positives if positives else [],
                    'source_file': a.get('source_file', ''),
                    'blob_url': generate_blob_url(a.get('source_file', '')),
                    'tags': ['va_checkin', a.get('overall_risk_level', 'medium').lower()],
                    'created_at': now_iso,
                    'updated_at': now_iso
                }
                documents.append(doc)
            
            # Upload to main index
            search_client = SearchClient(
                endpoint=AZURE_SEARCH_ENDPOINT,
                index_name=AZURE_SEARCH_INDEX,
                credential=credential
            )
            
            # Upload in batches
            batch_size = 100
            for i in range(0, len(documents), batch_size):
                batch = documents[i:i + batch_size]
                result = search_client.merge_or_upload_documents(batch)
                logging.info(f"   📤 Uploaded batch {i//batch_size + 1} ({len(batch)} docs) to main index")
            
            logging.info(f"   ✅ Main Copilot index updated ({len(documents)} meeting analyses)")
        
        # =====================================================================
        # Create/Update Coach Performance Index
        # =====================================================================
        coach_stats = data.get('coach_stats', {})
        coach_meetings = data.get('coach_meetings', {})
        
        if coach_stats:
            # Create coach index if not exists
            coach_fields = [
                SimpleField(name="id", type=SearchFieldDataType.String, key=True, filterable=True),
                SimpleField(name="document_type", type=SearchFieldDataType.String, filterable=True, facetable=True),
                SearchableField(name="coach_name", type=SearchFieldDataType.String, filterable=True, facetable=True, sortable=True),
                SearchableField(name="va_name", type=SearchFieldDataType.String, filterable=True, facetable=True),
                SearchableField(name="client_name", type=SearchFieldDataType.String, filterable=True, facetable=True),
                SimpleField(name="meeting_id", type=SearchFieldDataType.String, filterable=True),
                SimpleField(name="meeting_date", type=SearchFieldDataType.String, filterable=True, sortable=True),
                SimpleField(name="risk_level", type=SearchFieldDataType.String, filterable=True, facetable=True),
                SearchableField(name="evidence", type=SearchFieldDataType.String, analyzer_name="en.microsoft"),
                SimpleField(name="total_meetings", type=SearchFieldDataType.Int32, filterable=True, sortable=True),
                SimpleField(name="vas_coached", type=SearchFieldDataType.Int32, filterable=True, sortable=True),
                SimpleField(name="critical_cases", type=SearchFieldDataType.Int32, filterable=True, sortable=True),
                SimpleField(name="high_cases", type=SearchFieldDataType.Int32, filterable=True, sortable=True),
                SimpleField(name="escalations_handled", type=SearchFieldDataType.Int32, filterable=True, sortable=True),
                SimpleField(name="professionalism_score", type=SearchFieldDataType.Int32, filterable=True, sortable=True),
                SearchableField(name="performance_rating", type=SearchFieldDataType.String, filterable=True, facetable=True),
                SimpleField(name="updated_at", type=SearchFieldDataType.DateTimeOffset, filterable=True, sortable=True),
                SearchField(name="tags", type=SearchFieldDataType.Collection(SearchFieldDataType.String), searchable=True, filterable=True),
            ]
            
            coach_semantic = SemanticConfiguration(
                name="coach-semantic-config",
                prioritized_fields=SemanticPrioritizedFields(
                    title_field=SemanticField(field_name="coach_name"),
                    content_fields=[SemanticField(field_name="evidence")],
                    keywords_fields=[SemanticField(field_name="va_name"), SemanticField(field_name="performance_rating")]
                )
            )
            
            coach_index = SearchIndex(
                name=AZURE_SEARCH_COACH_INDEX,
                fields=coach_fields,
                semantic_search=SemanticSearch(configurations=[coach_semantic])
            )
            
            index_client.create_or_update_index(coach_index)
            logging.info(f"   ✅ Coach performance index '{AZURE_SEARCH_COACH_INDEX}' created/updated")
            
            # Upload coach summary documents
            coach_documents = []
            now_iso = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
            
            for coach_name, stats in coach_stats.items():
                total_mentions = stats['positive'] + stats['negative']
                prof_score = int((stats['positive'] / total_mentions) * 100) if total_mentions > 0 else 50
                rating = 'EXCELLENT' if prof_score >= 80 and stats['escalations'] > 0 else ('GOOD' if prof_score >= 60 else 'NEEDS REVIEW')
                
                # Summary document
                coach_documents.append({
                    'id': f"coach-summary-{coach_name.replace(' ', '-').lower()}",
                    'document_type': 'coach_summary',
                    'coach_name': coach_name,
                    'va_name': '',
                    'client_name': '',
                    'meeting_id': '',
                    'meeting_date': '',
                    'risk_level': '',
                    'evidence': f"{coach_name} has coached {len(stats['vas'])} VAs across {stats['total']} meetings. Performance: {rating}",
                    'total_meetings': stats['total'],
                    'vas_coached': len(stats['vas']),
                    'critical_cases': stats['critical'],
                    'high_cases': stats['high'],
                    'escalations_handled': stats['escalations'],
                    'professionalism_score': prof_score,
                    'performance_rating': rating,
                    'updated_at': now_iso,
                    'tags': ['coach', 'summary', rating.lower()]
                })
            
            # Upload meeting-level coach documents
            for coach_name, meetings in coach_meetings.items():
                for m in meetings:
                    coach_documents.append({
                        'id': f"coach-meeting-{m['meeting_id']}",
                        'document_type': 'coach_meeting',
                        'coach_name': coach_name,
                        'va_name': m['va'],
                        'client_name': m['client'],
                        'meeting_id': m['meeting_id'],
                        'meeting_date': m['date'],
                        'risk_level': m['risk'],
                        'evidence': m['evidence'],
                        'total_meetings': 0,
                        'vas_coached': 0,
                        'critical_cases': 0,
                        'high_cases': 0,
                        'escalations_handled': 0,
                        'professionalism_score': 0,
                        'performance_rating': '',
                        'updated_at': now_iso,
                        'tags': ['coach', 'meeting', m['risk'].lower()]
                    })
            
            # Upload to coach index
            coach_search_client = SearchClient(
                endpoint=AZURE_SEARCH_ENDPOINT,
                index_name=AZURE_SEARCH_COACH_INDEX,
                credential=credential
            )
            
            for i in range(0, len(coach_documents), batch_size):
                batch = coach_documents[i:i + batch_size]
                coach_search_client.merge_or_upload_documents(batch)
            
            logging.info(f"   ✅ Coach index updated ({len(coach_documents)} documents)")
        
        logging.info("   ✅ All Copilot indexes updated successfully")
        
    except ImportError as e:
        logging.warning(f"   ⚠️ Azure Search SDK not available: {e}")
    except Exception as e:
        logging.error(f"   ❌ Search index update failed: {e}")


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main(mytimer: func.TimerRequest) -> None:
    """Azure Function entry point."""
    utc_timestamp = datetime.utcnow().isoformat()
    
    if mytimer.past_due:
        logging.info('Timer is past due!')
    
    logging.info('=' * 60)
    logging.info('🚀 DAILY VA CHECK-IN PIPELINE STARTED')
    logging.info('=' * 60)
    logging.info(f'   Time: {utc_timestamp}')
    
    try:
        # Step 1: Download new transcripts
        downloaded = download_new_transcripts(logging, DAYS_TO_LOOK_BACK)
        
        # Step 2: Analyze check-in transcripts
        if downloaded:
            analyzed = analyze_new_transcripts(downloaded, logging)
        else:
            analyzed = []
            logging.info("   No new transcripts to analyze")
        
        # Step 3: Generate and upload CSVs (returns data for indexing)
        index_data = generate_and_upload_csvs(logging)
        
        # Step 4: Update Azure AI Search indexes for Copilot Agent
        if index_data:
            update_copilot_search_index(index_data, logging)
        
        # Step 5: Update state
        blob_service = get_blob_service()
        state = load_state(blob_service)
        state['last_run'] = utc_timestamp
        state['runs'] = state.get('runs', [])
        state['runs'].append({
            'time': utc_timestamp,
            'downloaded': len(downloaded),
            'analyzed': len(analyzed),
            'indexed': True
        })
        save_state(blob_service, state)
        
        logging.info('=' * 60)
        logging.info('✅ PIPELINE COMPLETED')
        logging.info('=' * 60)
        logging.info(f'   Downloaded: {len(downloaded)} transcripts')
        logging.info(f'   Analyzed: {len(analyzed)} check-ins')
        logging.info(f'   Indexed: Copilot + Coach indexes updated')
        
    except Exception as e:
        logging.error(f'❌ Pipeline failed: {e}')
        raise
