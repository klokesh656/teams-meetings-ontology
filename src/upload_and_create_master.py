"""
Upload transcripts to Azure Blob Storage and create master Excel file.
Creates two containers:
1. 'transcripts' - for VTT transcript files
2. 'reports' - for the master Excel metadata file
"""
import os
import re
import time
from datetime import datetime, timedelta
from dotenv import load_dotenv
import requests
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions
from azure.identity import ClientSecretCredential
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

load_dotenv()

# Azure credentials
CLIENT_ID = os.getenv('AZURE_CLIENT_ID')
CLIENT_SECRET = os.getenv('AZURE_CLIENT_SECRET')
TENANT_ID = os.getenv('AZURE_TENANT_ID')
STORAGE_CONNECTION_STRING = os.getenv('AZURE_STORAGE_CONNECTION_STRING')

# HR user ID
USER_ID = '81835016-79d5-4a15-91b1-c104e2cd9adb'

# Local directories
TRANSCRIPTS_DIR = 'transcripts'
OUTPUT_DIR = 'output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Blob containers
TRANSCRIPTS_CONTAINER = 'transcripts'
REPORTS_CONTAINER = 'reports'

# Rate limiting
REQUEST_DELAY = 0.3

class BlobUploader:
    def __init__(self):
        self.blob_service = BlobServiceClient.from_connection_string(STORAGE_CONNECTION_STRING)
        self._ensure_containers()
        
        # Parse account details from connection string
        parts = dict(item.split('=', 1) for item in STORAGE_CONNECTION_STRING.split(';') if '=' in item)
        self.account_name = parts.get('AccountName', '')
        self.account_key = parts.get('AccountKey', '')
    
    def _ensure_containers(self):
        """Create containers if they don't exist"""
        for container_name in [TRANSCRIPTS_CONTAINER, REPORTS_CONTAINER]:
            try:
                self.blob_service.create_container(container_name)
                print(f"Created container: {container_name}")
            except Exception as e:
                if 'ContainerAlreadyExists' in str(e):
                    pass
                else:
                    print(f"Container {container_name}: {e}")
    
    def upload_file(self, local_path: str, container: str, blob_name: str) -> str:
        """Upload file and return blob URL with SAS token"""
        container_client = self.blob_service.get_container_client(container)
        blob_client = container_client.get_blob_client(blob_name)
        
        with open(local_path, 'rb') as f:
            blob_client.upload_blob(f, overwrite=True)
        
        # Generate SAS URL valid for 1 year
        sas_token = generate_blob_sas(
            account_name=self.account_name,
            container_name=container,
            blob_name=blob_name,
            account_key=self.account_key,
            permission=BlobSasPermissions(read=True),
            expiry=datetime.utcnow() + timedelta(days=365)
        )
        
        return f"https://{self.account_name}.blob.core.windows.net/{container}/{blob_name}?{sas_token}"
    
    def upload_dataframe_as_excel(self, df: pd.DataFrame, container: str, blob_name: str) -> str:
        """Upload DataFrame as Excel file and return blob URL"""
        # Save locally first
        local_path = os.path.join(OUTPUT_DIR, blob_name)
        
        # Create styled Excel
        create_styled_excel(df, local_path)
        
        # Upload to blob
        return self.upload_file(local_path, container, blob_name)


def create_styled_excel(df: pd.DataFrame, filepath: str):
    """Create a styled Excel file from DataFrame"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Meeting Transcripts"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Cell styles
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
    
    # Set column widths
    column_widths = {
        'Meeting ID': 15,
        'Subject': 40,
        'Date': 12,
        'Time': 10,
        'Duration (min)': 12,
        'Organizer': 25,
        'Participants': 40,
        'Team': 20,
        'Transcript File': 30,
        'Blob Storage Link': 60,
        'Has Transcript': 12,
        'Sentiment Score': 12,
        'Churn Risk': 12,
        'Opportunity Score': 12,
        'Execution Reliability': 14,
        'Operational Complexity': 14,
        'Events': 40,
        'Summary': 60,
        'Key Concerns': 40,
        'Key Positives': 40,
        'Action Items': 40,
        'Analyzed At': 18
    }
    
    for col_idx, col_name in enumerate(df.columns, 1):
        width = column_widths.get(col_name, 20)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(filepath)
    print(f"Saved Excel to: {filepath}")


def get_graph_token():
    """Get Microsoft Graph API token"""
    credential = ClientSecretCredential(TENANT_ID, CLIENT_ID, CLIENT_SECRET)
    return credential.get_token('https://graph.microsoft.com/.default').token


def get_all_transcripts_metadata(headers):
    """Get all transcripts with their meeting metadata"""
    url = f"https://graph.microsoft.com/beta/users/{USER_ID}/onlineMeetings/getAllTranscripts(meetingOrganizerUserId='{USER_ID}')"
    
    all_transcripts = []
    while url:
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                all_transcripts.extend(data.get('value', []))
                url = data.get('@odata.nextLink')
                time.sleep(REQUEST_DELAY)
            else:
                print(f"Error fetching transcripts: {resp.status_code}")
                break
        except Exception as e:
            print(f"Error: {e}")
            time.sleep(5)
    
    return all_transcripts


def get_meeting_details(headers, meeting_id):
    """Get meeting details"""
    url = f"https://graph.microsoft.com/beta/users/{USER_ID}/onlineMeetings/{meeting_id}"
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        time.sleep(REQUEST_DELAY)
        if resp.status_code == 200:
            return resp.json()
    except:
        pass
    return {}


def parse_vtt_filename(filename):
    """Parse date and subject from VTT filename"""
    # Format: YYYYMMDD_HHMMSS_Subject.vtt or unknown_Subject.vtt
    match = re.match(r'(\d{8})_(\d{6})_(.+)\.vtt', filename)
    if match:
        date_str = match.group(1)
        time_str = match.group(2)
        subject = match.group(3).replace('_', ' ')
        try:
            date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
            time = f"{time_str[:2]}:{time_str[2:4]}:{time_str[4:6]}"
        except:
            date = 'Unknown'
            time = 'Unknown'
        return date, time, subject
    
    # unknown_Subject.vtt
    match = re.match(r'unknown_(.+)\.vtt', filename)
    if match:
        subject = match.group(1).replace('_', ' ')
        return 'Unknown', 'Unknown', subject
    
    return 'Unknown', 'Unknown', filename.replace('.vtt', '')


def main():
    print("=" * 70)
    print("UPLOAD TRANSCRIPTS & CREATE MASTER EXCEL")
    print("=" * 70)
    
    # Initialize
    uploader = BlobUploader()
    token = get_graph_token()
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    
    # Get list of local VTT files
    vtt_files = [f for f in os.listdir(TRANSCRIPTS_DIR) if f.endswith('.vtt')]
    print(f"\nFound {len(vtt_files)} VTT files to process")
    
    # Get transcript metadata from API
    print("\nFetching transcript metadata from API...")
    api_transcripts = get_all_transcripts_metadata(headers)
    print(f"Found {len(api_transcripts)} transcripts in API")
    
    # Build meeting details cache
    print("\nFetching meeting details (this may take a while)...")
    meeting_cache = {}
    
    # Process each VTT file
    records = []
    uploaded = 0
    
    for i, filename in enumerate(vtt_files):
        date, time_str, subject = parse_vtt_filename(filename)
        local_path = os.path.join(TRANSCRIPTS_DIR, filename)
        
        print(f"\n[{i+1}/{len(vtt_files)}] {filename}")
        
        # Upload to blob storage
        try:
            blob_url = uploader.upload_file(
                local_path, 
                TRANSCRIPTS_CONTAINER, 
                filename
            )
            uploaded += 1
            print(f"   ✅ Uploaded to blob storage")
        except Exception as e:
            blob_url = f"Error: {str(e)[:50]}"
            print(f"   ❌ Upload failed: {e}")
        
        # Try to find matching meeting in API data
        meeting_info = {}
        for t in api_transcripts:
            meeting_id = t.get('meetingId', '')
            if meeting_id and meeting_id not in meeting_cache:
                meeting_cache[meeting_id] = get_meeting_details(headers, meeting_id)
            
            cached = meeting_cache.get(meeting_id, {})
            api_subject = cached.get('subject', '')
            
            # Match by subject
            if api_subject and subject.lower() in api_subject.lower():
                meeting_info = cached
                break
        
        # Get participants
        participants = []
        if meeting_info.get('participants', {}).get('attendees'):
            for att in meeting_info['participants']['attendees']:
                identity = att.get('identity', {}).get('user', {})
                name = identity.get('displayName', '')
                if name:
                    participants.append(name)
        
        # Calculate duration from file size (rough estimate)
        file_size = os.path.getsize(local_path)
        # Rough estimate: 1KB ≈ 1 minute of transcript
        estimated_duration = max(1, file_size // 1000)
        
        # Create record
        record = {
            'Meeting ID': meeting_info.get('id', '')[:20] + '...' if meeting_info.get('id') else '',
            'Subject': subject,
            'Date': date,
            'Time': time_str,
            'Duration (min)': meeting_info.get('endDateTime', '') and estimated_duration or '',
            'Organizer': meeting_info.get('participants', {}).get('organizer', {}).get('identity', {}).get('user', {}).get('displayName', 'HR@Our-Assistants'),
            'Participants': ', '.join(participants[:5]) if participants else '',
            'Team': '',  # To be filled manually or via mapping
            'Transcript File': filename,
            'Blob Storage Link': blob_url,
            'Has Transcript': 'Yes',
            # Analysis columns (to be filled by analyze_transcripts.py)
            'Sentiment Score': None,
            'Churn Risk': None,
            'Opportunity Score': None,
            'Execution Reliability': None,
            'Operational Complexity': None,
            'Events': None,
            'Summary': None,
            'Key Concerns': None,
            'Key Positives': None,
            'Action Items': None,
            'Analyzed At': None
        }
        records.append(record)
    
    # Create DataFrame
    df = pd.DataFrame(records)
    
    # Sort by date (newest first)
    df['_sort_date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('_sort_date', ascending=False, na_position='last')
    df = df.drop('_sort_date', axis=1)
    
    print("\n" + "=" * 70)
    print("CREATING MASTER EXCEL")
    print("=" * 70)
    
    # Save locally
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_filename = f"meeting_transcripts_master_{timestamp}.xlsx"
    local_excel = os.path.join(OUTPUT_DIR, excel_filename)
    create_styled_excel(df, local_excel)
    
    # Upload Excel to reports container
    print("\nUploading Excel to Azure Blob Storage (reports container)...")
    try:
        excel_blob_url = uploader.upload_file(
            local_excel,
            REPORTS_CONTAINER,
            excel_filename
        )
        print(f"✅ Excel uploaded to: {excel_blob_url[:80]}...")
    except Exception as e:
        print(f"❌ Excel upload failed: {e}")
        excel_blob_url = None
    
    # Also save a "latest" version
    try:
        latest_excel = os.path.join(OUTPUT_DIR, 'meeting_transcripts_latest.xlsx')
        create_styled_excel(df, latest_excel)
        uploader.upload_file(latest_excel, REPORTS_CONTAINER, 'meeting_transcripts_latest.xlsx')
        print("✅ Also uploaded as 'meeting_transcripts_latest.xlsx'")
    except Exception as e:
        print(f"Note: Could not create latest version: {e}")
    
    # Print summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Total VTT files: {len(vtt_files)}")
    print(f"Uploaded to blob: {uploaded}")
    print(f"Records in Excel: {len(df)}")
    print(f"\nLocal Excel: {os.path.abspath(local_excel)}")
    print(f"\nBlob Storage Containers:")
    print(f"  - Transcripts: https://aidevelopement.blob.core.windows.net/{TRANSCRIPTS_CONTAINER}/")
    print(f"  - Reports: https://aidevelopement.blob.core.windows.net/{REPORTS_CONTAINER}/")
    
    if excel_blob_url:
        print(f"\nMaster Excel URL (valid for 1 year):")
        print(f"  {excel_blob_url}")


if __name__ == '__main__':
    main()
