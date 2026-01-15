#!/usr/bin/env python3
"""
Generate CSV files for Power Automate and Power BI from batch analysis results.

Outputs:
1. va_risk_summary.csv - Summary view for Power BI dashboards
2. pending_suggestions.csv - All suggestions pending review (for Power Automate emails)
3. critical_alerts.csv - Critical/High risk cases for immediate stakeholder action
4. va_signals_detail.csv - Detailed signal data for deep analysis
5. stakeholder_review.csv - Formatted for email notifications via Power Automate
6. client_feedback_form.xlsx - Excel file with table for Power Automate
7. coach_performance_summary.csv - Coach/HR performance metrics
"""

import json
import csv
import os
import hashlib
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Azure Storage Configuration
STORAGE_CONNECTION_STRING = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
OUTPUT_CONTAINER = "pipeline-outputs"

# SharePoint Configuration - BT Group plc site
SHAREPOINT_SITE_ID = os.getenv('SHAREPOINT_SITE_ID', 'bt2685218p1.sharepoint.com,8470aef2-afc9-4665-8730-e63c87b0ebac,66af7ccb-d9d6-41da-b647-1c62bc4fd46e')
SHAREPOINT_DRIVE_ID = os.getenv('SHAREPOINT_DRIVE_ID', 'b!8q5whMmvZUaHMOY8h7DrrMt8r2bW2dpBtkccYrxP1G6UufIavqAUT4u5gAnnEB1q')
SHAREPOINT_FOLDER = os.getenv('SHAREPOINT_FOLDER', 'Client Feedback')

# Azure AD for Graph API (SharePoint)
TENANT_ID = os.getenv('AZURE_TENANT_ID')
CLIENT_ID = os.getenv('AZURE_CLIENT_ID')
CLIENT_SECRET = os.getenv('AZURE_CLIENT_SECRET')

# Paths
OUTPUT_DIR = Path(__file__).parent.parent / "output"
BATCH_RESULTS = OUTPUT_DIR / "batch_analysis_results_20260107_011511.json"
ANALYSIS_HISTORY = OUTPUT_DIR / "checkin_analysis_history.json"
PENDING_SUGGESTIONS = OUTPUT_DIR / "pending_suggestions.json"


def get_blob_service():
    """Get Azure Blob Storage service client."""
    from azure.storage.blob import BlobServiceClient
    return BlobServiceClient.from_connection_string(STORAGE_CONNECTION_STRING)


def upload_to_azure_blob(filepath, blob_name=None, container=OUTPUT_CONTAINER):
    """Upload file to Azure Blob Storage."""
    if not STORAGE_CONNECTION_STRING:
        print(f"   ⚠️ Skipping Azure upload (no connection string)")
        return None
    
    try:
        blob_service = get_blob_service()
        container_client = blob_service.get_container_client(container)
        
        # Create container if it doesn't exist
        try:
            container_client.create_container()
        except:
            pass
        
        blob_name = blob_name or filepath.name
        blob_path = f"csv/{datetime.now().strftime('%Y%m%d')}/{blob_name}"
        blob_client = container_client.get_blob_client(blob_path)
        
        with open(filepath, 'rb') as f:
            blob_client.upload_blob(f, overwrite=True)
        
        # Also upload to 'latest' folder
        latest_blob = container_client.get_blob_client(f"latest/{blob_name}")
        with open(filepath, 'rb') as f:
            latest_blob.upload_blob(f, overwrite=True)
        
        print(f"   ☁️ Uploaded to Azure: {blob_path}")
        return blob_path
    except Exception as e:
        print(f"   ⚠️ Azure upload failed: {e}")
        return None


def upload_to_sharepoint(filepath, folder_path=None):
    """Upload file to SharePoint using Microsoft Graph API with app permissions.
    
    Requires SharePoint site ID and drive ID to be configured for app-only auth.
    """
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
        print(f"   ⚠️ Skipping SharePoint upload (missing credentials)")
        return None
    
    if not SHAREPOINT_SITE_ID:
        print(f"   ⚠️ Skipping SharePoint upload (SHAREPOINT_SITE_ID not configured)")
        print(f"   💡 To enable SharePoint upload, add to .env:")
        print(f"      SHAREPOINT_SITE_ID=your-site-id")
        print(f"      SHAREPOINT_DRIVE_ID=your-drive-id (optional)")
        return None
    
    try:
        import requests
        
        # Get access token
        token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
        token_data = {
            'grant_type': 'client_credentials',
            'client_id': CLIENT_ID,
            'client_secret': CLIENT_SECRET,
            'scope': 'https://graph.microsoft.com/.default'
        }
        token_response = requests.post(token_url, data=token_data)
        access_token = token_response.json().get('access_token')
        
        if not access_token:
            print(f"   ⚠️ Failed to get SharePoint access token")
            return None
        
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/octet-stream'
        }
        
        folder = folder_path or SHAREPOINT_FOLDER
        filename = filepath.name
        
        # Use site-specific endpoint for app-only auth
        if SHAREPOINT_DRIVE_ID:
            # Use specific drive
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}/root:/{folder}/{filename}:/content"
        else:
            # Use default document library
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drive/root:/{folder}/{filename}:/content"
        
        with open(filepath, 'rb') as f:
            response = requests.put(upload_url, headers=headers, data=f)
        
        if response.status_code in [200, 201]:
            web_url = response.json().get('webUrl', '')
            print(f"   ✅ Uploaded to SharePoint: {folder}/{filename}")
            print(f"   🔗 URL: {web_url}")
            return web_url
        else:
            error_msg = response.text[:300] if response.text else 'Unknown error'
            print(f"   ⚠️ SharePoint upload returned {response.status_code}")
            print(f"      Error: {error_msg}")
            return None
            
    except Exception as e:
        print(f"   ⚠️ SharePoint upload failed: {e}")
        return None


def generate_meeting_id(source_file, va_name='', meeting_date=''):
    """Generate a consistent Meeting ID based on source file.
    
    This creates a unique, stable ID that remains the same across pipeline runs,
    enabling Power BI relationships and Power Automate status tracking.
    """
    if source_file:
        file_hash = hashlib.md5(source_file.encode()).hexdigest()[:6].upper()
        return f"MTG-{file_hash}"
    elif va_name and meeting_date:
        combined = f"{va_name}_{meeting_date}"
        file_hash = hashlib.md5(combined.encode()).hexdigest()[:6].upper()
        return f"MTG-{file_hash}"
    return "MTG-UNKNOWN"


def load_json(filepath):
    """Load JSON file safely."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


def generate_va_risk_summary():
    """Generate VA risk summary for Power BI dashboard."""
    print("\n📊 Generating VA Risk Summary...")
    
    data = load_json(BATCH_RESULTS)
    if not data:
        return
    
    # Aggregate by VA
    va_stats = {}
    for record in data.get('processed', []):
        va = record['va_name']
        if va not in va_stats:
            va_stats[va] = {
                'va_name': va,
                'total_checkins': 0,
                'critical_count': 0,
                'high_count': 0,
                'medium_count': 0,
                'low_count': 0,
                'total_signals': 0,
                'total_suggestions': 0,
                'latest_date': record['date'],
                'latest_risk': record['risk_level'],
                'clients': set()
            }
        
        stats = va_stats[va]
        stats['total_checkins'] += 1
        stats['total_signals'] += record['signals_count']
        stats['total_suggestions'] += record['suggestions_count']
        
        if record['client_name'] and record['client_name'] != 'Unknown':
            stats['clients'].add(record['client_name'][:50])  # Truncate long names
        
        # Count by risk level
        risk = record['risk_level'].lower()
        if risk == 'critical':
            stats['critical_count'] += 1
        elif risk == 'high':
            stats['high_count'] += 1
        elif risk == 'medium':
            stats['medium_count'] += 1
        else:
            stats['low_count'] += 1
        
        # Track latest
        if record['date'] >= stats['latest_date']:
            stats['latest_date'] = record['date']
            stats['latest_risk'] = record['risk_level']
    
    # Calculate risk scores and write CSV
    output_file = OUTPUT_DIR / "va_risk_summary.csv"
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'VA Name', 'Total Check-ins', 'Critical Count', 'High Count', 
            'Medium Count', 'Low Count', 'Total Signals', 'Total Suggestions',
            'Risk Score', 'Current Risk Level', 'Latest Check-in Date', 
            'Known Clients', 'Attention Required'
        ])
        
        for va, stats in sorted(va_stats.items(), key=lambda x: (
            x[1]['critical_count'] * 100 + x[1]['high_count'] * 10
        ), reverse=True):
            # Calculate risk score (weighted)
            risk_score = (
                stats['critical_count'] * 100 +
                stats['high_count'] * 25 +
                stats['medium_count'] * 5 +
                stats['low_count'] * 1
            )
            
            attention = 'CRITICAL' if stats['critical_count'] > 0 else \
                       'HIGH' if stats['high_count'] > 0 else \
                       'MONITOR' if stats['medium_count'] > 2 else 'OK'
            
            writer.writerow([
                stats['va_name'],
                stats['total_checkins'],
                stats['critical_count'],
                stats['high_count'],
                stats['medium_count'],
                stats['low_count'],
                stats['total_signals'],
                stats['total_suggestions'],
                risk_score,
                stats['latest_risk'].upper(),
                stats['latest_date'],
                '; '.join(list(stats['clients'])[:3]) if stats['clients'] else 'Unknown',
                attention
            ])
    
    print(f"   ✅ Saved: {output_file}")
    return len(va_stats)


def generate_pending_suggestions_csv():
    """Generate pending suggestions for Power Automate workflow."""
    print("\n📋 Generating Pending Suggestions CSV...")
    
    data = load_json(PENDING_SUGGESTIONS)
    if not data:
        # Try to extract from analysis history
        data = load_json(ANALYSIS_HISTORY)
        if not data:
            print("   ⚠️ No pending suggestions found")
            return 0
    
    output_file = OUTPUT_DIR / "pending_suggestions_review.csv"
    suggestion_count = 0
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Suggestion ID', 'VA Name', 'Client', 'Meeting Date', 'Risk Level',
            'Issue', 'Category', 'Urgency', 'Suggestion', 'Rationale',
            'Status', 'Review Link'
        ])
        
        if isinstance(data, dict) and 'suggestions' in data:
            suggestions = data['suggestions']
        elif isinstance(data, list):
            suggestions = data
        else:
            suggestions = []
        
        for sugg in suggestions:
            if isinstance(sugg, dict):
                writer.writerow([
                    sugg.get('id', f'SUGG-{suggestion_count+1:04d}'),
                    sugg.get('va_name', 'Unknown'),
                    sugg.get('client_name', 'Unknown'),
                    sugg.get('meeting_date', ''),
                    sugg.get('risk_level', 'medium').upper(),
                    sugg.get('issue', '')[:200],
                    sugg.get('category', ''),
                    sugg.get('urgency', 'monitor'),
                    sugg.get('suggestion', '')[:500],
                    sugg.get('rationale', '')[:300],
                    sugg.get('status', 'pending'),
                    f"Review in Outlier Insights Engine"
                ])
                suggestion_count += 1
    
    print(f"   ✅ Saved: {output_file} ({suggestion_count} suggestions)")
    return suggestion_count


def generate_critical_alerts():
    """Generate critical alerts for immediate stakeholder attention."""
    print("\n🚨 Generating Critical Alerts CSV...")
    
    data = load_json(BATCH_RESULTS)
    if not data:
        return 0
    
    # Also load detailed analysis
    history = load_json(ANALYSIS_HISTORY) or {'analyses': []}
    
    # Create lookup for detailed info
    detail_lookup = {}
    for analysis in history.get('analyses', []):
        key = f"{analysis.get('va_name', '')}_{analysis.get('meeting_date', '')}"
        detail_lookup[key] = analysis
    
    output_file = OUTPUT_DIR / "critical_alerts.csv"
    alert_count = 0
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Meeting ID', 'Priority', 'VA Name', 'Client', 'Meeting Date', 'Risk Level',
            'Signals Count', 'Executive Summary', 'Key Findings',
            'Escalation Reason', 'Immediate Actions Required',
            'Stakeholder', 'Review Status'
        ])
        
        for record in data.get('processed', []):
            risk = record['risk_level'].lower()
            if risk in ['critical', 'high']:
                key = f"{record['va_name']}_{record['date']}"
                detail = detail_lookup.get(key, {})
                
                # Generate Meeting ID
                source_file = detail.get('source_file', '')
                meeting_id = generate_meeting_id(source_file, record['va_name'], record['date'])
                
                # Determine priority
                priority = 'P1 - CRITICAL' if risk == 'critical' else 'P2 - HIGH'
                
                # Get summary and findings
                summary = detail.get('executive_summary', 'Analysis pending review')[:500]
                findings = '; '.join(detail.get('key_findings', [])[:3])[:400]
                escalation = detail.get('escalation_reason', '')[:300]
                
                # Get immediate actions from suggestions
                actions = []
                for sugg in detail.get('ai_suggestions', [])[:2]:
                    if sugg.get('urgency') in ['immediate', 'within_48h']:
                        actions.append(sugg.get('suggestion', '')[:150])
                
                writer.writerow([
                    meeting_id,
                    priority,
                    record['va_name'],
                    record['client_name'][:50] if record['client_name'] else 'Unknown',
                    record['date'],
                    risk.upper(),
                    record['signals_count'],
                    summary,
                    findings,
                    escalation,
                    ' | '.join(actions) if actions else 'Review detailed report',
                    'Pending Assignment',
                    'PENDING REVIEW'
                ])
                alert_count += 1
    
    print(f"   ✅ Saved: {output_file} ({alert_count} alerts)")
    return alert_count


def generate_signals_detail():
    """Generate detailed signals data for Power BI analysis."""
    print("\n📈 Generating Signals Detail CSV...")
    
    history = load_json(ANALYSIS_HISTORY)
    if not history:
        return 0
    
    output_file = OUTPUT_DIR / "va_signals_detail.csv"
    signal_count = 0
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Meeting ID', 'Analysis ID', 'VA Name', 'Client', 'Meeting Date', 'Overall Risk',
            'Signal ID', 'Signal Category', 'Evidence', 'Confidence',
            'VA Status', 'Client Health', 'Source File', 'Transcript Blob Link'
        ])
        
        for analysis in history.get('analyses', []):
            source_file = analysis.get('source_file', '')
            meeting_id = generate_meeting_id(
                source_file, 
                analysis.get('va_name', ''), 
                analysis.get('meeting_date', '')
            )
            blob_link = analysis.get('transcript_blob_link', '')
            
            for signal in analysis.get('detected_signals', []):
                # Determine signal category from ID
                signal_id = signal.get('signal_id', '')
                if signal_id.startswith('VA'):
                    category = 'VA Issue'
                elif signal_id.startswith('CL'):
                    category = 'Client Issue'
                elif signal_id.startswith('RH'):
                    category = 'Relationship/HR Issue'
                else:
                    category = 'Other'
                
                writer.writerow([
                    meeting_id,
                    analysis.get('analysis_id', ''),
                    analysis.get('va_name', 'Unknown'),
                    analysis.get('client_name', 'Unknown')[:50],
                    analysis.get('meeting_date', ''),
                    analysis.get('overall_risk_level', 'unknown').upper(),
                    signal_id,
                    category,
                    signal.get('evidence', '')[:300],
                    signal.get('confidence', 'medium'),
                    analysis.get('va_status', ''),
                    analysis.get('client_health', ''),
                    source_file,
                    blob_link
                ])
                signal_count += 1
    
    print(f"   ✅ Saved: {output_file} ({signal_count} signals)")
    return signal_count


def generate_stakeholder_review():
    """Generate stakeholder review file for Power Automate email workflow."""
    print("\n📧 Generating Stakeholder Review CSV...")
    
    data = load_json(BATCH_RESULTS)
    history = load_json(ANALYSIS_HISTORY) or {'analyses': []}
    
    if not data:
        return 0
    
    # Create lookup
    detail_lookup = {}
    for analysis in history.get('analyses', []):
        key = f"{analysis.get('va_name', '')}_{analysis.get('meeting_date', '')}"
        detail_lookup[key] = analysis
    
    output_file = OUTPUT_DIR / "stakeholder_review.csv"
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Review ID', 'Date Generated', 'VA Name', 'Client', 'Check-in Date',
            'Risk Level', 'Risk Score', 'Signals Detected', 'Suggestions Count',
            'Summary', 'Top Recommendation', 'Stakeholder Email', 'Review URL',
            'Approval Status', 'Approved By', 'Approval Date', 'Notes'
        ])
        
        review_count = 0
        for record in data.get('processed', []):
            key = f"{record['va_name']}_{record['date']}"
            detail = detail_lookup.get(key, {})
            
            # Calculate risk score
            risk_weights = {'critical': 100, 'high': 75, 'medium': 25, 'low': 5}
            risk_score = risk_weights.get(record['risk_level'].lower(), 10)
            
            # Get top recommendation
            top_rec = ''
            for sugg in detail.get('ai_suggestions', []):
                if sugg.get('urgency') == 'immediate':
                    top_rec = sugg.get('suggestion', '')[:200]
                    break
            if not top_rec and detail.get('ai_suggestions'):
                top_rec = detail.get('ai_suggestions', [{}])[0].get('suggestion', '')[:200]
            
            review_id = f"REV-{datetime.now().strftime('%Y%m%d')}-{review_count+1:04d}"
            
            writer.writerow([
                review_id,
                datetime.now().strftime('%Y-%m-%d %H:%M'),
                record['va_name'],
                record['client_name'][:50] if record['client_name'] else 'Unknown',
                record['date'],
                record['risk_level'].upper(),
                risk_score,
                record['signals_count'],
                record['suggestions_count'],
                detail.get('executive_summary', '')[:300],
                top_rec,
                '',  # Stakeholder email - to be filled in Power Automate
                '',  # Review URL - to be filled with SharePoint link
                'PENDING',
                '',  # Approved by
                '',  # Approval date
                ''   # Notes
            ])
            review_count += 1
    
    print(f"   ✅ Saved: {output_file} ({review_count} reviews)")
    return review_count


def generate_kpi_summary():
    """Generate KPI summary for Power BI executive dashboard."""
    print("\n📊 Generating KPI Summary CSV...")
    
    data = load_json(BATCH_RESULTS)
    history = load_json(ANALYSIS_HISTORY) or {'analyses': []}
    
    if not data:
        return
    
    # Calculate KPIs
    processed = data.get('processed', [])
    total_analyses = len(processed)
    
    risk_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    total_signals = 0
    total_suggestions = 0
    unique_vas = set()
    unique_clients = set()
    escalations_needed = 0
    
    for record in processed:
        risk_counts[record['risk_level'].lower()] += 1
        total_signals += record['signals_count']
        total_suggestions += record['suggestions_count']
        unique_vas.add(record['va_name'])
        if record['client_name'] and record['client_name'] != 'Unknown':
            unique_clients.add(record['client_name'][:30])
    
    # Count escalations from history
    for analysis in history.get('analyses', []):
        if analysis.get('escalation_needed'):
            escalations_needed += 1
    
    output_file = OUTPUT_DIR / "kpi_dashboard_summary.csv"
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['KPI', 'Value', 'Category', 'Trend', 'Target', 'Status'])
        
        kpis = [
            ('Total Check-ins Analyzed', total_analyses, 'Volume', '↑', '-', 'INFO'),
            ('Unique VAs Monitored', len(unique_vas), 'Coverage', '→', '-', 'INFO'),
            ('Unique Clients', len(unique_clients), 'Coverage', '→', '-', 'INFO'),
            ('Critical Risk Cases', risk_counts['critical'], 'Risk', '↓', '0', 
             'CRITICAL' if risk_counts['critical'] > 0 else 'OK'),
            ('High Risk Cases', risk_counts['high'], 'Risk', '↓', '< 5', 
             'WARNING' if risk_counts['high'] > 5 else 'OK'),
            ('Medium Risk Cases', risk_counts['medium'], 'Risk', '→', '-', 'MONITOR'),
            ('Low Risk Cases', risk_counts['low'], 'Risk', '↑', '> 50%', 'INFO'),
            ('Total Signals Detected', total_signals, 'Analysis', '→', '-', 'INFO'),
            ('Total Suggestions Generated', total_suggestions, 'Action', '→', '-', 'INFO'),
            ('Escalations Required', escalations_needed, 'Urgency', '↓', '0', 
             'CRITICAL' if escalations_needed > 10 else 'WARNING' if escalations_needed > 5 else 'OK'),
            ('Critical Risk Rate (%)', round(risk_counts['critical']/total_analyses*100, 1) if total_analyses else 0, 
             'Risk', '↓', '< 5%', 'CRITICAL' if risk_counts['critical']/total_analyses*100 > 5 else 'OK'),
            ('High Risk Rate (%)', round(risk_counts['high']/total_analyses*100, 1) if total_analyses else 0, 
             'Risk', '↓', '< 15%', 'WARNING' if risk_counts['high']/total_analyses*100 > 15 else 'OK'),
            ('Avg Signals per Check-in', round(total_signals/total_analyses, 1) if total_analyses else 0, 
             'Quality', '→', '-', 'INFO'),
            ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Meta', '-', '-', 'INFO'),
        ]
        
        for kpi in kpis:
            writer.writerow(kpi)
    
    print(f"   ✅ Saved: {output_file}")


def generate_all_in_one():
    """Generate a single comprehensive CSV with all check-in data."""
    print("\n📦 Generating All-in-One Dataset CSV...")
    
    data = load_json(BATCH_RESULTS)
    history = load_json(ANALYSIS_HISTORY) or {'analyses': []}
    
    if not data:
        return
    
    # Create detail lookup
    detail_lookup = {}
    for analysis in history.get('analyses', []):
        key = f"{analysis.get('va_name', '')}_{analysis.get('meeting_date', '')}"
        detail_lookup[key] = analysis
    
    output_file = OUTPUT_DIR / "checkin_analysis_complete.csv"
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'VA Name', 'Client', 'Meeting Date', 'Risk Level', 'Risk Score',
            'VA Status', 'Client Health', 'Signals Count', 'Suggestions Count',
            'Signal IDs', 'Key Finding 1', 'Key Finding 2', 'Key Finding 3',
            'Top Suggestion', 'Escalation Needed', 'Escalation Reason',
            'Executive Summary', 'Positive Indicators', 'Analysis ID', 'Analyzed At'
        ])
        
        for record in data.get('processed', []):
            key = f"{record['va_name']}_{record['date']}"
            detail = detail_lookup.get(key, {})
            
            # Risk score
            risk_weights = {'critical': 100, 'high': 75, 'medium': 25, 'low': 5}
            risk_score = risk_weights.get(record['risk_level'].lower(), 10)
            
            # Signal IDs
            signal_ids = [s.get('signal_id', '') for s in detail.get('detected_signals', [])]
            
            # Key findings
            findings = detail.get('key_findings', ['', '', ''])
            while len(findings) < 3:
                findings.append('')
            
            # Top suggestion
            top_sugg = detail.get('ai_suggestions', [{}])[0].get('suggestion', '')[:250] if detail.get('ai_suggestions') else ''
            
            # Positive indicators
            positives = '; '.join(detail.get('positive_indicators', [])[:2])[:200]
            
            writer.writerow([
                record['va_name'],
                record['client_name'][:50] if record['client_name'] else 'Unknown',
                record['date'],
                record['risk_level'].upper(),
                risk_score,
                detail.get('va_status', ''),
                detail.get('client_health', ''),
                record['signals_count'],
                record['suggestions_count'],
                ', '.join(signal_ids),
                findings[0][:150] if findings[0] else '',
                findings[1][:150] if len(findings) > 1 and findings[1] else '',
                findings[2][:150] if len(findings) > 2 and findings[2] else '',
                top_sugg,
                'Yes' if detail.get('escalation_needed') else 'No',
                detail.get('escalation_reason', '')[:200],
                detail.get('executive_summary', '')[:300],
                positives,
                detail.get('analysis_id', ''),
                detail.get('analyzed_at', '')
            ])
    
    print(f"   ✅ Saved: {output_file}")


def generate_client_feedback_form():
    """Generate client feedback form as XLSX with Excel Table for Power Automate.
    
    Creates an Excel file with a proper Table that Power Automate can use
    to trigger flows when rows are modified.
    """
    print("\n📝 Generating Client Feedback Form XLSX...")
    
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("   ⚠️ openpyxl/pandas not installed. Run: pip install openpyxl pandas")
        return 0
    
    data = load_json(BATCH_RESULTS)
    history = load_json(ANALYSIS_HISTORY) or {'analyses': []}
    
    if not data:
        return 0
    
    # Create lookup for detailed info
    detail_lookup = {}
    for analysis in history.get('analyses', []):
        key = f"{analysis.get('va_name', '')}_{analysis.get('meeting_date', '')}"
        detail_lookup[key] = analysis
    
    # Prepare data rows
    rows = []
    headers = [
        'Meeting_ID', 'VA_Name', 'Client', 'Meeting_Date', 'Risk_Level',
        'Issue_Summary', 'AI_Recommendation',
        'Client_Agrees', 'Client_Priority', 'Client_Suggestion', 
        'Client_Action_Taken', 'Client_Notes', 'Reviewer_Name', 'Review_Date',
        'Feedback_Status'
    ]
    
    for record in data.get('processed', []):
        risk = record['risk_level'].lower()
        if risk in ['critical', 'high', 'medium']:
            key = f"{record['va_name']}_{record['date']}"
            detail = detail_lookup.get(key, {})
            
            source_file = detail.get('source_file', '')
            meeting_id = generate_meeting_id(source_file, record['va_name'], record['date'])
            
            summary = detail.get('executive_summary', '')[:300]
            top_suggestion = ''
            for sugg in detail.get('ai_suggestions', [])[:1]:
                top_suggestion = sugg.get('suggestion', '')[:300]
            
            rows.append({
                'Meeting_ID': meeting_id,
                'VA_Name': record['va_name'],
                'Client': record['client_name'][:50] if record['client_name'] else 'Unknown',
                'Meeting_Date': record['date'],
                'Risk_Level': risk.upper(),
                'Issue_Summary': summary,
                'AI_Recommendation': top_suggestion,
                'Client_Agrees': '',
                'Client_Priority': '',
                'Client_Suggestion': '',
                'Client_Action_Taken': '',
                'Client_Notes': '',
                'Reviewer_Name': '',
                'Review_Date': '',
                'Feedback_Status': 'PENDING'
            })
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Create Excel workbook with table
    output_file = OUTPUT_DIR / "client_feedback_form.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ClientFeedback"
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_num, row_data in enumerate(rows, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
    
    # Create Excel Table
    if len(rows) > 0:
        table_ref = f"A1:{chr(64 + len(headers))}{len(rows) + 1}"
        table = Table(displayName="ClientFeedbackTable", ref=table_ref)
        
        # Add style
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    
    # Adjust column widths
    column_widths = {
        'A': 12, 'B': 15, 'C': 20, 'D': 12, 'E': 10,
        'F': 40, 'G': 40, 'H': 12, 'I': 15, 'J': 30,
        'K': 25, 'L': 25, 'M': 15, 'N': 12, 'O': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(output_file)
    
    print(f"   ✅ Saved: {output_file} ({len(rows)} items)")
    print(f"   📊 Excel Table 'ClientFeedbackTable' created for Power Automate")
    
    # Also save CSV version for backup
    csv_file = OUTPUT_DIR / "client_feedback_form.csv"
    df.to_csv(csv_file, index=False)
    
    return len(rows), output_file


def generate_coach_performance():
    """Generate coach/HR performance analysis CSV.
    
    Analyzes coach behavior based on transcript evidence to evaluate:
    - Communication quality
    - Professionalism
    - Empathy and support
    - Follow-through on action items
    - Coaching effectiveness
    """
    print("\n👔 Generating Coach Performance CSV...")
    
    history = load_json(ANALYSIS_HISTORY)
    if not history:
        return 0
    
    # Known coaches/HR (can be expanded)
    known_coaches = ['Louise', 'Shey', 'Shey Geraldes']
    
    # Track coach metrics
    coach_stats = {}
    coach_meetings = {}  # Store meeting details per coach
    
    for analysis in history.get('analyses', []):
        # Extract coach name from evidence in signals
        coach_name = None
        coach_evidence = []
        
        for signal in analysis.get('detected_signals', []):
            evidence = signal.get('evidence', '')
            for coach in known_coaches:
                if coach.lower() in evidence.lower():
                    # Normalize coach name
                    if 'shey' in coach.lower():
                        coach_name = 'Shey Geraldes'
                    else:
                        coach_name = coach
                    coach_evidence.append(evidence)
                    break
        
        # Also check positive indicators and findings for coach mentions
        for finding in analysis.get('key_findings', []):
            for coach in known_coaches:
                if coach.lower() in finding.lower():
                    if 'shey' in coach.lower():
                        coach_name = 'Shey Geraldes'
                    else:
                        coach_name = coach
                    coach_evidence.append(finding)
        
        for positive in analysis.get('positive_indicators', []):
            for coach in known_coaches:
                if coach.lower() in positive.lower():
                    if 'shey' in coach.lower():
                        coach_name = 'Shey Geraldes'
                    else:
                        coach_name = coach
                    coach_evidence.append(f"[POSITIVE] {positive}")
        
        if coach_name:
            if coach_name not in coach_stats:
                coach_stats[coach_name] = {
                    'total_meetings': 0,
                    'critical_cases': 0,
                    'high_cases': 0,
                    'medium_cases': 0,
                    'low_cases': 0,
                    'escalations_handled': 0,
                    'positive_mentions': 0,
                    'negative_mentions': 0,
                    'vas_coached': set()
                }
                coach_meetings[coach_name] = []
            
            stats = coach_stats[coach_name]
            stats['total_meetings'] += 1
            stats['vas_coached'].add(analysis.get('va_name', 'Unknown'))
            
            risk = analysis.get('overall_risk_level', 'low').lower()
            if risk == 'critical':
                stats['critical_cases'] += 1
            elif risk == 'high':
                stats['high_cases'] += 1
            elif risk == 'medium':
                stats['medium_cases'] += 1
            else:
                stats['low_cases'] += 1
            
            if analysis.get('escalation_needed'):
                stats['escalations_handled'] += 1
            
            # Count positive vs negative mentions
            for ev in coach_evidence:
                if '[POSITIVE]' in ev or 'support' in ev.lower() or 'help' in ev.lower():
                    stats['positive_mentions'] += 1
                elif 'issue' in ev.lower() or 'problem' in ev.lower() or 'miss' in ev.lower():
                    stats['negative_mentions'] += 1
            
            # Store meeting detail
            source_file = analysis.get('source_file', '')
            meeting_id = generate_meeting_id(source_file, analysis.get('va_name', ''), analysis.get('meeting_date', ''))
            coach_meetings[coach_name].append({
                'meeting_id': meeting_id,
                'va_name': analysis.get('va_name', 'Unknown'),
                'client': analysis.get('client_name', 'Unknown'),
                'date': analysis.get('meeting_date', ''),
                'risk': risk.upper(),
                'evidence': ' | '.join(coach_evidence[:2])[:400]
            })
    
    # Generate summary CSV
    output_file = OUTPUT_DIR / "coach_performance_summary.csv"
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Coach Name', 'Total Meetings', 'VAs Coached', 
            'Critical Cases', 'High Cases', 'Medium Cases', 'Low Cases',
            'Escalations Handled', 'Positive Mentions', 'Negative Mentions',
            'Professionalism Score', 'Performance Rating'
        ])
        
        for coach, stats in coach_stats.items():
            # Calculate professionalism score (0-100)
            total_mentions = stats['positive_mentions'] + stats['negative_mentions']
            if total_mentions > 0:
                prof_score = int((stats['positive_mentions'] / total_mentions) * 100)
            else:
                prof_score = 50  # Neutral if no mentions
            
            # Performance rating based on metrics
            if prof_score >= 80 and stats['escalations_handled'] > 0:
                rating = 'EXCELLENT'
            elif prof_score >= 60:
                rating = 'GOOD'
            elif prof_score >= 40:
                rating = 'NEEDS IMPROVEMENT'
            else:
                rating = 'REQUIRES REVIEW'
            
            writer.writerow([
                coach,
                stats['total_meetings'],
                len(stats['vas_coached']),
                stats['critical_cases'],
                stats['high_cases'],
                stats['medium_cases'],
                stats['low_cases'],
                stats['escalations_handled'],
                stats['positive_mentions'],
                stats['negative_mentions'],
                prof_score,
                rating
            ])
    
    print(f"   ✅ Saved: {output_file} ({len(coach_stats)} coaches)")
    
    # Generate detailed coach meetings CSV
    detail_file = OUTPUT_DIR / "coach_performance_detail.csv"
    detail_count = 0
    with open(detail_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Meeting ID', 'Coach Name', 'VA Name', 'Client', 'Meeting Date',
            'Risk Level', 'Evidence/Observations', 'Leadership Review Notes'
        ])
        
        for coach, meetings in coach_meetings.items():
            for m in meetings:
                writer.writerow([
                    m['meeting_id'],
                    coach,
                    m['va_name'],
                    m['client'][:50],
                    m['date'],
                    m['risk'],
                    m['evidence'],
                    ''  # For leadership to fill
                ])
                detail_count += 1
    
    print(f"   ✅ Saved: {detail_file} ({detail_count} meeting records)")
    return len(coach_stats), output_file, detail_file


def upload_all_to_azure():
    """Upload all output files to Azure Blob Storage."""
    print("\n" + "=" * 70)
    print("☁️ UPLOADING TO AZURE BLOB STORAGE")
    print("=" * 70)
    
    files_to_upload = [
        "va_risk_summary.csv",
        "va_signals_detail.csv", 
        "critical_alerts.csv",
        "kpi_dashboard_summary.csv",
        "all_meetings_detail.csv",
        "checkin_analysis_complete.csv",
        "stakeholder_review.csv",
        "pending_suggestions_review.csv",
        "client_feedback_form.xlsx",
        "client_feedback_form.csv",
        "coach_performance_summary.csv",
        "coach_performance_detail.csv"
    ]
    
    uploaded = 0
    for filename in files_to_upload:
        filepath = OUTPUT_DIR / filename
        if filepath.exists():
            result = upload_to_azure_blob(filepath)
            if result:
                uploaded += 1
    
    print(f"\n   📊 Uploaded {uploaded}/{len(files_to_upload)} files to Azure")
    return uploaded


def upload_feedback_to_sharepoint():
    """Upload client feedback form to SharePoint BT group documents."""
    print("\n" + "=" * 70)
    print("📤 UPLOADING CLIENT FEEDBACK TO SHAREPOINT")
    print("=" * 70)
    
    xlsx_file = OUTPUT_DIR / "client_feedback_form.xlsx"
    if xlsx_file.exists():
        result = upload_to_sharepoint(xlsx_file, SHAREPOINT_FOLDER)
        if result:
            print(f"   ✅ Client feedback form available at: {result}")
            return result
    
    print("   ⚠️ client_feedback_form.xlsx not found")
    return None


def main():
    """Generate all CSV files and upload to Azure/SharePoint."""
    print("=" * 70)
    print("🚀 GENERATING CSV/XLSX FILES FOR POWER AUTOMATE & POWER BI")
    print("=" * 70)
    print(f"   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Generate all CSVs
    va_count = generate_va_risk_summary()
    generate_pending_suggestions_csv()
    alert_count = generate_critical_alerts()
    signal_count = generate_signals_detail()
    review_count = generate_stakeholder_review()
    generate_kpi_summary()
    generate_all_in_one()
    
    # Client feedback (XLSX with Excel Table) and coach performance
    feedback_result = generate_client_feedback_form()
    generate_coach_performance()
    
    # Upload to Azure Blob Storage
    upload_all_to_azure()
    
    # Upload client feedback to SharePoint
    sharepoint_url = upload_feedback_to_sharepoint()
    
    print("\n" + "=" * 70)
    print("✅ GENERATION & UPLOAD COMPLETE")
    print("=" * 70)
    print(f"""
📁 Output Files Created in {OUTPUT_DIR}:

   FOR POWER BI DASHBOARDS:
   ├── va_risk_summary.csv            - VA risk scores & trends
   ├── va_signals_detail.csv          - Signal-level analysis (with Meeting ID)
   ├── critical_alerts.csv            - P1/P2 alerts (with Meeting ID)
   ├── kpi_dashboard_summary.csv      - Executive KPIs
   └── checkin_analysis_complete.csv  - Full dataset

   FOR POWER AUTOMATE EMAIL WORKFLOW:
   ├── stakeholder_review.csv         - Review items with approval tracking
   └── pending_suggestions_review.csv - All suggestions pending review

   FOR CLIENT FEEDBACK (XLSX WITH EXCEL TABLE):
   └── client_feedback_form.xlsx      - Excel Table for Power Automate triggers
       📍 Uploaded to SharePoint: {SHAREPOINT_FOLDER}

   FOR COACH/HR PERFORMANCE (LEADERSHIP):
   ├── coach_performance_summary.csv  - Coach metrics & professionalism scores
   └── coach_performance_detail.csv   - Meeting-level coach analysis

☁️ AZURE BLOB STORAGE:
   Container: {OUTPUT_CONTAINER}
   Path: csv/YYYYMMDD/ and latest/

📤 SHAREPOINT:
   Folder: {SHAREPOINT_FOLDER}
   File: client_feedback_form.xlsx

📊 POWER AUTOMATE SETUP:
   1. Add SharePoint connector
   2. Trigger: "When a row is modified" on ClientFeedbackTable
   3. Get Meeting_ID from modified row
   4. Send notification/update Power BI dataset

📊 POWER BI RELATIONSHIPS:
   all_meetings_detail ─── Meeting ID ───> va_signals_detail
   all_meetings_detail ─── Meeting ID ───> critical_alerts  
   all_meetings_detail ─── Meeting ID ───> client_feedback_form
   coach_performance_detail ─── Meeting ID ───> all_meetings_detail
    """)


if __name__ == "__main__":
    main()
